"""
Civil Estimates Backend — FastAPI + YOLO
Detects doors, walls, and windows from uploaded floor plan images.
Serves the frontend static files as well.
"""

# ── Fix PyTorch DLL loading on Windows ───────────────────────────────────
import os
import sys
import platform

if platform.system() == "Windows":
    # Workaround for WinError 1114 with PyTorch on some Windows setups
    torch_lib = os.path.join(
        os.path.dirname(sys.executable), "Lib", "site-packages", "torch", "lib"
    )
    if os.path.isdir(torch_lib):
        os.add_dll_directory(torch_lib)
    # Also try user-local site-packages
    for sp in sys.path:
        candidate = os.path.join(sp, "torch", "lib")
        if os.path.isdir(candidate):
            try:
                os.add_dll_directory(candidate)
            except OSError:
                pass

import io
import base64
from pathlib import Path

import cv2
import numpy as np

import fitz  # PyMuPDF — for PDF-to-image conversion

from fastapi import FastAPI, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.requests import Request
from starlette.responses import FileResponse
from PIL import Image
from ultralytics import YOLO

from wall_detection import detect_walls

# ── App setup ────────────────────────────────────────────────────────────────
app = FastAPI(title="Civil Estimates API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Load YOLO model ─────────────────────────────────────────────────────────
MODEL_PATH = Path(__file__).resolve().parent.parent / "best.pt"
model = YOLO(str(MODEL_PATH))

# ── Default scale factor (pixels per foot) — no OCR dependency ──────────────
DEFAULT_PIXELS_PER_FOOT = 20

# Class name mapping from the model
CLASS_NAMES = model.names  # e.g. {0: 'd1', 1: 'd2', 2: 'w'}

# Human-readable label mapping
LABEL_MAP = {
    "d1": "Single Door",
    "d2": "Double Door",
    "w":  "Window",
}




# ── Per-class confidence thresholds ──────────────────────────────────────
# Class IDs: 0 = d1 (Single Door), 1 = d2 (Double Door), 2 = w (Window)
CLASS_CONF_THRESHOLDS = {
    0: 0.12,   # Single Door — reject low-conf false positives
    1: 0.05,   # Double Door — keep even low-conf to avoid missing
    2: 0.10,   # Window
}
DEFAULT_CONF = 0.10

# ── Per-class max bounding-box size (pixels) ─────────────────────────────
CLASS_MAX_PIXELS = {
    0: (400, 400),   # Single Door max width x height
    1: (500, 500),   # Double Door
    2: (600, 200),   # Window — typically wide and narrow
}
DEFAULT_MAX_PIXELS = (600, 600)


def run_inference(image: Image.Image, filter_class: str | None = None):
    """Run YOLO inference on a PIL image and return structured results.

    Uses optimised settings discovered through experimentation:
    - imgsz=1536 for higher-resolution feature extraction
    - agnostic_nms=True to merge cross-class overlapping boxes
    - Per-class confidence thresholds to cut false positives
    - Per-class bounding-box size limits to reject structural artefacts
    """
    results = model.predict(
        source=image,
        conf=0.05,           # low global conf — we filter per-class below
        iou=0.5,
        agnostic_nms=True,   # merge overlapping d1/d2 boxes
        imgsz=1536,          # higher res → better window + door recall
        verbose=False,
    )
    result = results[0]

    # ── Per-class filtering (confidence + size) ──────────────────────────
    valid_indices = []
    detections = []

    for i, box in enumerate(result.boxes):
        cls_id = int(box.cls[0])
        raw_label = CLASS_NAMES.get(cls_id, f"class_{cls_id}")
        label = LABEL_MAP.get(raw_label, raw_label)
        confidence = float(box.conf[0])
        x1, y1, x2, y2 = [float(c) for c in box.xyxy[0]]
        w, h = x2 - x1, y2 - y1

        # Per-class confidence threshold
        min_conf = CLASS_CONF_THRESHOLDS.get(cls_id, DEFAULT_CONF)
        # Per-class max bounding-box size
        max_w, max_h = CLASS_MAX_PIXELS.get(cls_id, DEFAULT_MAX_PIXELS)

        if confidence >= min_conf and w <= max_w and h <= max_h:
            valid_indices.append(i)
            detections.append({
                "id": len(detections) + 1,
                "label": label,
                "confidence": round(confidence, 4),
                "bbox": {"x1": round(x1, 1), "y1": round(y1, 1),
                          "x2": round(x2, 1), "y2": round(y2, 1)},
                "width": round(w, 1),
                "height": round(h, 1),
            })

    # Keep only valid boxes in the YOLO result object (for plot)
    filtered_result = result[valid_indices]

    # ── Optional class filter (user picks "Single Door" etc.) ────────────
    if filter_class:
        detections = [d for d in detections
                      if d["label"].lower() == filter_class.lower()]
        keep_indices = []
        for i, box in enumerate(filtered_result.boxes):
            cls_id = int(box.cls[0])
            raw = CLASS_NAMES.get(cls_id, "")
            lbl = LABEL_MAP.get(raw, raw)
            if lbl.lower() == filter_class.lower():
                keep_indices.append(i)
        plot_result = filtered_result[keep_indices]
    else:
        plot_result = filtered_result

    # ── Summary counts ───────────────────────────────────────────────────
    summary = {}
    for d in detections:
        summary[d["label"]] = summary.get(d["label"], 0) + 1

    # ── Annotated image (base64) — custom numbered boxes ────────────────
    # Draw bounding boxes with numbered labels instead of YOLO's default
    # class+confidence text.  The frontend renders an HTML legend panel.
    img_array = np.array(image)  # RGB
    annotated = img_array.copy()

    # Colour palette per label type (BGR for OpenCV)
    BOX_COLORS = {
        "single door": (230, 180, 50),    # cyan-ish
        "double door":  (180, 120, 230),  # purple-ish
        "window":       (100, 210, 140),  # green-ish
    }
    DEFAULT_COLOR = (200, 200, 200)

    for det in detections:
        bb = det["bbox"]
        x1, y1, x2, y2 = int(bb["x1"]), int(bb["y1"]), int(bb["x2"]), int(bb["y2"])
        label_key = det["label"].lower()
        color = BOX_COLORS.get(label_key, DEFAULT_COLOR)

        # Draw bounding box
        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 4)

        # Draw numbered tag near top-left corner of the box
        tag = str(det["id"])
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.9
        thickness = 3
        (tw, th), baseline = cv2.getTextSize(tag, font, font_scale, thickness)
        tag_pad = 6
        # Tag background rectangle
        cv2.rectangle(
            annotated,
            (x1, y1 - th - tag_pad * 2),
            (x1 + tw + tag_pad * 2, y1),
            color, -1,  # filled
        )
        # Tag text (white on coloured background)
        cv2.putText(
            annotated, tag,
            (x1 + tag_pad, y1 - tag_pad),
            font, font_scale, (255, 255, 255), thickness, cv2.LINE_AA,
        )

    annotated_img = Image.fromarray(annotated)
    buf = io.BytesIO()
    annotated_img.save(buf, format="PNG")
    annotated_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

    return {
        "annotated_image": annotated_b64,
        "detections": detections,
        "summary": summary,
        "total_objects": len(detections),
    }


# ── Helpers ──────────────────────────────────────────────────────────────────
def pdf_to_image(pdf_bytes: bytes, page_num: int = 0, dpi: int = 300) -> Image.Image:
    """Convert a PDF page to a PIL Image at ~300 DPI.

    Simple fast render — no text removal processing.
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc.load_page(page_num)

    # Render to image at 2x zoom (~300 DPI)
    mat = fitz.Matrix(2, 2)
    pix = page.get_pixmap(matrix=mat)
    img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
    if pix.n == 4:
        img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
    doc.close()

    img = Image.fromarray(img_array)
    return img


# ── API Endpoints ────────────────────────────────────────────────────────────
@app.post("/api/detect")
async def detect_objects(
    file: UploadFile = File(...),
    object_type: str | None = Query(None, description="Filter by object type"),
    page: int = Query(0, description="PDF page number (0-indexed)"),
):
    """Upload a floor plan image or PDF and get detection results.

    Runs YOLO object detection, then wall detection (region-growing).
    When object_type='Wall', the walls_image is returned as the annotated image.
    """
    contents = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".pdf"):
        image = pdf_to_image(contents, page_num=page)
    else:
        image = Image.open(io.BytesIO(contents)).convert("RGB")

    # ── YOLO inference ───────────────────────────────────────────────
    yolo_filter = object_type if (object_type and object_type.lower() != "wall") else None
    results = run_inference(image, filter_class=yolo_filter)

    # ── Extract window and door bounding boxes from YOLO detections ──
    window_boxes = []
    door_boxes = []
    for det in results.get("detections", []):
        bb = det["bbox"]
        box = (int(bb["x1"]), int(bb["y1"]), int(bb["x2"]), int(bb["y2"]))
        label_lower = det["label"].lower()
        if label_lower == "window":
            window_boxes.append(box)
        elif "door" in label_lower:
            door_boxes.append(box)

    # ── Wall detection (region-growing) ──────────────────────────────
    wall_results = detect_walls(image, window_boxes=window_boxes, door_boxes=door_boxes)

    # ── Merge wall results into the response ─────────────────────────
    results.update(wall_results)
    results["scale"] = {
        "pixels_per_foot": DEFAULT_PIXELS_PER_FOOT,
        "source": "default",
    }

    # When the user filters by "Wall", show the walls image instead
    if object_type and object_type.lower() == "wall":
        results["annotated_image"] = wall_results["walls_image"]

    return JSONResponse(content=results)


@app.get("/api/classes")
async def get_classes():
    """Return the list of detectable object classes."""
    return {"classes": list(CLASS_NAMES.values())}


@app.get("/api/health")
async def health():
    return {"status": "ok", "model_loaded": MODEL_PATH.exists()}




# ── Excel Export: persistent Detection_Record.xlsx ───────────────────────────
EXCEL_PATH = Path(__file__).resolve().parent.parent / "Detection_Record.xlsx"

@app.post("/api/export-excel")
async def export_excel(request: Request):
    """Append detection results to the persistent Detection_Record.xlsx file.

    Expected JSON body:
    {
        "file_name": "floor_plan.jpg",
        "single_door_detected": 2, "single_door_actual": 3,
        "double_door_detected": 1, "double_door_actual": 1,
        "window_detected": 4, "window_actual": 5,
        "wall_count": 8,
        "floor_coverage_pct": 34.2
    }
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    body = await request.json()

    file_name = body.get("file_name", "Unknown")
    sd_det = body.get("single_door_detected", 0)
    sd_act = body.get("single_door_actual", 0)
    dd_det = body.get("double_door_detected", 0)
    dd_act = body.get("double_door_actual", 0)
    w_det  = body.get("window_detected", 0)
    w_act  = body.get("window_actual", 0)
    wall_count = body.get("wall_count", "")
    floor_cov  = body.get("floor_coverage_pct", "")

    # Calculate totals
    total_det = sd_det + dd_det + w_det
    total_act = sd_act + dd_act + w_act
    total_diff = total_act - total_det
    total_acc = round((min(total_det, total_act) / total_act) * 100, 1) if total_act > 0 else 100

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Headers for the sheet ──
    HEADERS = [
        "S.No", "Date & Time", "File Name",
        "Single Door (Detected)", "Single Door (Actual)", "Single Door (Diff)",
        "Double Door (Detected)", "Double Door (Actual)", "Double Door (Diff)",
        "Window (Detected)", "Window (Actual)", "Window (Diff)",
        "Total Detected", "Total Actual", "Total Diff", "Accuracy %",
        "Walls Detected", "Floor Coverage %",
    ]

    # ── Open or create workbook ──
    if EXCEL_PATH.exists():
        wb = load_workbook(str(EXCEL_PATH))
        ws = wb.active
        row_num = ws.max_row + 1
        serial = row_num - 1  # subtract header row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Record"

        # Style the header row
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Set column widths
        widths = [6, 20, 25, 18, 16, 16, 18, 16, 16, 14, 14, 12, 14, 12, 10, 12, 16, 16]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else None].width = w
        # Handle columns A-R properly
        from openpyxl.utils import get_column_letter
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

        row_num = 2
        serial = 1

    # ── Append the data row ──
    row_data = [
        serial,
        timestamp,
        file_name,
        sd_det, sd_act, sd_act - sd_det,
        dd_det, dd_act, dd_act - dd_det,
        w_det,  w_act,  w_act - w_det,
        total_det, total_act, total_diff, f"{total_acc}%",
        wall_count, f"{floor_cov}%" if floor_cov else "",
    ]

    # Style data row
    from openpyxl.styles import Alignment as Align, Border as Brd, Side as Sd
    data_align = Align(horizontal="center", vertical="center")
    thin = Brd(
        left=Sd(style="thin"), right=Sd(style="thin"),
        top=Sd(style="thin"), bottom=Sd(style="thin"),
    )

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = data_align
        cell.border = thin

    # ── Save ──
    wb.save(str(EXCEL_PATH))

    return {
        "status": "ok",
        "message": f"Result #{serial} saved to Detection_Record.xlsx",
        "row_number": serial,
        "file_path": str(EXCEL_PATH),
    }


# ── Serve Frontend ───────────────────────────────────────────────────────────
FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"

app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


@app.get("/")
async def serve_frontend():
    return FileResponse(str(FRONTEND_DIR / "index.html"))
